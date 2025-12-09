
import os
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False

def generate_report(df_block: pd.DataFrame, df_details: pd.DataFrame, output_dir: str, timestamp: str):
    """
    生成分析报告：包含图表和Excel文件
    """
    print("\n正在生成可视化报告...")
    
    # 1. 生成图表
    chart_path = _generate_charts(df_block, output_dir, timestamp)
    
    # 2. 生成Excel
    excel_path = os.path.join(output_dir, "global_analysis_report.xlsx")
    
    # 准备数据：为详情表添加板块排名信息，以便排序
    # 创建板块排名映射
    block_rank_map = {row['细分板块']: i for i, row in df_block.reset_index().iterrows()}
    
    # 复制一份详情数据进行处理
    df_sorted_details = df_details.copy()
    # 添加板块排名列
    df_sorted_details['板块排名'] = df_sorted_details['Block'].map(block_rank_map)
    # 排序：先按板块排名，再按个股成交额降序
    df_sorted_details = df_sorted_details.sort_values(['板块排名', '成交额'], ascending=[True, False])
    # 移除辅助列
    df_sorted_details = df_sorted_details.drop(columns=['板块排名'])
    
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # 写入各个Sheet
            df_block.to_excel(writer, sheet_name='板块概览', index=False)
            df_sorted_details.to_excel(writer, sheet_name='个股详情', index=False)
            
        # 3. 后处理Excel：调整格式、插入图表
        _format_excel(excel_path, chart_path)
        
        print(f"✅ 报告已生成: {excel_path}")
        
    except Exception as e:
        print(f"❌ 生成报告失败: {e}")

def _generate_charts(df_block: pd.DataFrame, output_dir: str, timestamp: str) -> str:
    """生成板块涨跌幅柱状图"""
    # 取前20个板块（如果太多的话）
    plot_data = df_block.head(20).sort_values('成交额加权涨跌幅(%)', ascending=True)
    
    plt.figure(figsize=(12, 8))
    
    # 颜色映射：涨红跌绿
    colors = ['red' if x >= 0 else 'green' for x in plot_data['成交额加权涨跌幅(%)']]
    
    bars = plt.barh(plot_data['细分板块'], plot_data['成交额加权涨跌幅(%)'], color=colors)
    
    plt.title(f'细分板块成交额加权涨跌幅排名 (Top 20) - {timestamp}', fontsize=14)
    plt.xlabel('涨跌幅 (%)', fontsize=12)
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    
    # 添加数值标签
    for bar in bars:
        width = bar.get_width()
        label_x_pos = width + 0.1 if width >= 0 else width - 0.1
        ha = 'left' if width >= 0 else 'right'
        plt.text(label_x_pos, bar.get_y() + bar.get_height()/2, f'{width:.2f}%', 
                 va='center', ha=ha, fontsize=10)
    
    plt.tight_layout()
    
    chart_path = os.path.join(output_dir, "block_ranking_chart.png")
    plt.savefig(chart_path, dpi=100)
    plt.close()
    
    return chart_path

def _format_excel(excel_path: str, chart_path: str):
    """格式化Excel并插入图表"""
    wb = load_workbook(excel_path)
    
    # 1. 格式化 "板块概览" Sheet
    ws_block = wb['板块概览']
    _adjust_column_width(ws_block)
    
    # 插入图表到板块概览Sheet的右侧
    if os.path.exists(chart_path):
        img = Image(chart_path)
        # 调整图片大小
        img.width = 800
        img.height = 500
        ws_block.add_image(img, 'G2') # 假设数据占用了A-E列，从G列开始放图
        
    # 2. 格式化 "个股详情" Sheet
    ws_details = wb['个股详情']
    _adjust_column_width(ws_details)
    
    wb.save(excel_path)

def _adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 50) # 限制最大宽度
